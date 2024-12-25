import sys
import os
import subprocess
from antlr4 import *
from openpyxl import load_workbook

# Define your variables
DIR = os.path.dirname(__file__)
ANTLR_JAR = 'C:/antlr/antlr4-4.9.2-complete.jar'  # your location is going here
CPL_Dest = 'CompiledFiles'
SRC = 'Sample.g4'
EXCEL_FILE = 'financial_calculations.xlsx'  # Name of your Excel file

def printUsage():
    print('python run.py gen')
    print('python run.py test')

def printBreak():
    print('-----------------------------------------------')

def generateAntlr2Python():
    print('Antlr4 is running...')
    subprocess.run(['java', '-jar', ANTLR_JAR, '-o', CPL_Dest, '-no-listener', '-visitor', '-Dlanguage=Python3', SRC])    
    print('Generate successfully')

def runCode(astTree):    
    from CodeRunner import CodeRunner
    code_runner = CodeRunner()
    result = astTree.accept(code_runner)
    
    print("Results:")
    for var, value in code_runner.variables.items():
        print(f"{var} = {value}")

def runTest():
    print('Running testcases...')
       
    from CompiledFiles.SampleLexer import SampleLexer
    from CompiledFiles.SampleParser import SampleParser
    from antlr4.error.ErrorListener import ErrorListener
    from ASTGeneration import ASTGeneration

    class CustomErrorListener(ErrorListener):
        def syntaxError(self, recognizer, offendingSymbol, line, column, msg, e):
            print(f"Input rejected: {msg}")
            exit(1)  # Exit the program with an error

    excel_file = os.path.join(DIR, "./tests", EXCEL_FILE)
    
    # Load the workbook and select the active sheet
    workbook = load_workbook(filename=excel_file)
    sheet = workbook.active

    # Generate input string from Excel data
    input_string = ""
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] and row[1]:  # Ensure both cells have values
            input_string += f"{row[0]} = {row[1]};\n"

    # Create input stream from the generated string
    input_stream = InputStream(input_string)
    lexer = SampleLexer(input_stream)
    stream = CommonTokenStream(lexer)
    parser = SampleParser(stream)   
    parser.removeErrorListeners()
    parser.addErrorListener(CustomErrorListener())    

    try:
        tree = parser.program()
        print("Input accepted")
    except SystemExit:        
        return
    
    printBreak()
    print('Parse tree:')
    print(tree.toStringTree(recog=parser))
    
    ast_generator = ASTGeneration()
    asttree = tree.accept(ast_generator)    
    print('AST:')
    print(asttree)
    
    runCode(asttree)

def main(argv):
    print('Complete jar file ANTLR  :  ' + str(ANTLR_JAR))
    print('Length of arguments      :  ' + str(len(argv)))    
    printBreak()

    if len(argv) < 1:
        printUsage()
    elif argv[0] == 'gen':
        generateAntlr2Python()    
    elif argv[0] == 'test':       
        runTest()
    else:
        printUsage()

if __name__ == '__main__':
    main(sys.argv[1:])

