import tkinter as tk
from tkinter import filedialog, messagebox
import sys
from antlr4 import *
from openpyxl import load_workbook

# Import your existing modules
from CompiledFiles.SampleLexer import SampleLexer
from CompiledFiles.SampleParser import SampleParser
from ASTGeneration import ASTGeneration
from CodeRunner import CodeRunner

class FinancialCalculator:
    def __init__(self, master):
        self.master = master
        master.title("Financial Calculator")
        master.geometry("400x500")
        master.configure(bg='#f0f0f0')

        self.create_widgets()

    def create_widgets(self):
        # File selection
        self.file_frame = tk.Frame(self.master, bg='#f0f0f0')
        self.file_frame.pack(pady=20)

        self.file_label = tk.Label(self.file_frame, text="Select Excel File:", bg='#f0f0f0')
        self.file_label.pack(side=tk.LEFT)

        self.file_path = tk.StringVar()
        self.file_entry = tk.Entry(self.file_frame, textvariable=self.file_path, width=30)
        self.file_entry.pack(side=tk.LEFT, padx=5)

        self.browse_button = tk.Button(self.file_frame, text="Browse", command=self.browse_file)
        self.browse_button.pack(side=tk.LEFT)

        # Calculate button
        self.calculate_button = tk.Button(self.master, text="Calculate", command=self.calculate)
        self.calculate_button.pack(pady=10)

        # Results display
        self.results_frame = tk.Frame(self.master, bg='#f0f0f0')
        self.results_frame.pack(pady=20, fill=tk.BOTH, expand=True)

        self.results_label = tk.Label(self.results_frame, text="Results:", bg='#f0f0f0')
        self.results_label.pack()

        self.results_text = tk.Text(self.results_frame, height=15, width=40)
        self.results_text.pack(padx=10, pady=5)

    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.file_path.set(filename)

    def calculate(self):
        file_path = self.file_path.get()
        if not file_path:
            messagebox.showerror("Error", "Please select an Excel file.")
            return

        try:
            results = self.process_excel(file_path)
            self.display_results(results)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def process_excel(self, filepath):
        workbook = load_workbook(filename=filepath)
        sheet = workbook.active

        # Generate input string from Excel data
        input_string = ""
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] and row[1]:  # Ensure both cells have values
                input_string += f"{row[0]} = {row[1]};\n"

        # Create input stream from the generated string
        input_stream = InputStream(input_string)
        lexer = SampleLexer(input_stream)
        stream = CommonTokenStream(lexer)
        parser = SampleParser(stream)

        tree = parser.program()
        ast_generator = ASTGeneration()
        asttree = tree.accept(ast_generator)

        code_runner = CodeRunner()
        asttree.accept(code_runner)

        return code_runner.variables

    def display_results(self, results):
        self.results_text.config(state='normal')       
        self.results_text.delete('1.0', tk.END)
        for key, value in results.items():
            self.results_text.insert(tk.END, f"{key:<{20}} = {value}\n")
        self.results_text.config(state='disabled')            

if __name__ == '__main__':
    root = tk.Tk()
    app = FinancialCalculator(root)
    root.mainloop()

